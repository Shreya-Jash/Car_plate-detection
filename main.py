import cv2
import pytesseract
import pandas as pd
import numpy as np
import openpyxl

from flask import Flask, render_template, request


# Function to process the uploaded image
import os
import openpyxl

app = Flask(__name__)

# Define the classifier to detect license plates
plate_cascade = cv2.CascadeClassifier('haarcascade_russian_plate_number.xml')


def process_image(file):
    # Load the image
    img = cv2.imdecode(np.fromstring(file.read(), np.uint8), cv2.IMREAD_COLOR)

    # Convert the image to grayscale
    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Apply Gaussian blur to the image
    img_gray = cv2.GaussianBlur(img_gray, (5, 5), 0)

    # Detect license plates in the image
    plates = plate_cascade.detectMultiScale(img_gray, 1.1, 4)

    # Create a list to store the results
    results = []

    # Loop over each detected plate
    for (x, y, w, h) in plates:
        # Crop the image to the plate region
        plate_roi = img_gray[y:y+h, x:x+w]
    
        # Apply thresholding to the plate region
        _, plate_thresh = cv2.threshold(plate_roi, 150, 255, cv2.THRESH_BINARY)
    
        # Perform OCR on the thresholded plate region
        plate_text = pytesseract.image_to_string(plate_thresh, lang='eng', config='--psm 8 --oem 3')
    
        # Append the result to the list
        results.append(plate_text)
    
    # Save the results to an Excel file using pandas
    df = pd.DataFrame({'License Plate': results})
    
    # Check if the file exists
    if os.path.isfile('output.xlsx'):
        # Load the existing workbook
        wb = openpyxl.load_workbook('output.xlsx')
        
        # Select the active worksheet
        ws = wb.active
        
        # Get the next available row
        next_row = ws.max_row + 1
        
        # Append the new results to the worksheet
        for result in results:
            ws.cell(row=next_row, column=1).value = result
            next_row += 1
        
        # Save the workbook
        wb.save('output.xlsx')
    else:
        # If the file doesn't exist, create a new workbook and save the results to it
        df.to_excel('output.xlsx', index=False, header=False)

    return results


# Route for the file upload form
@app.route('/')
def upload_form():
    return render_template('upload_form.html')

# Route for processing the uploaded image
@app.route('/process_form', methods=['POST'])
def process_form():
    file = request.files['image_file']
    results = process_image(file)
    return render_template('results.html', results=request.form['results'])

if __name__ == '__main__':
    app.run(debug=True)
